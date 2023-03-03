import os
from openpyxl import load_workbook
from PyPDF2 import PdfReader
import tabula
import numpy as np
import pandas as pd
from selenium.webdriver.cocliente_3n.action_chains import ActionChains
from selenium import webdriver
from selenium.webdriver.cocliente_3n.by import By
from time import sleep
import pyautogui
import cv2
import pytesseract
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import simpledialog



def formatar(texto):
    if str(texto).count(".") > 0 and str(texto).count(" ") == 0:
        texto = float((str(texto).replace(".", "")).replace(",", "."))
    return texto

def on_submit():
    global leitura_arquivos
    global vars
    for i, conta in enumerate(contas_menu):
        leitura = vars[i].get()
        leitura_arquivos[conta] = leitura
    root2.quit()

#dados planilha excel a ser editada
#selecionar planilha
root_planilha = tk.Tk()
root_planilha.withdraw() # Esconder a janela raiz
messagebox.showinfo("Planilha", "Escolha a planilha onde os dados serão salvos.")
planilha_path = filedialog.askopenfilename() # Mostrar a janela de seleção de pasta
planilha = load_workbook(planilha_path, keep_vba=True)
#abas
checks = planilha.worksheets[0]
santander = planilha.worksheets[2]
btg = planilha.worksheets[4]
itau = planilha.worksheets[6]
xp = planilha.worksheets[8]
class_rda_multi = planilha.worksheets[10] #fundo_2 - pdf
class_rda_ii = planilha.worksheets[13]
class_rda_prev = planilha.worksheets[16]
quantum = planilha.worksheets[36]


root_planilha.destroy()

#obter caminho de cada arquivo e organizar em lista
root = tk.Tk()
root.withdraw() # Esconder a janela raiz
messagebox.showinfo("Pasta do mês", "Selecione a pasta mensal que contém os arquivos a serem analisados.")
path_pasta_mes = filedialog.askdirectory() # Mostrar a janela de seleção de pasta
#nome das pastas no computador que roda o programa
path_pasta_rda_fi_multi = path_pasta_mes + '/Extrato RDA FI Multim/'
path_pasta_rda_ii = path_pasta_mes + '/Extrato fundo_3/'
path_pasta_rda_prev = path_pasta_mes + '/Extrato fundo_1/'
path_pasta_btg = path_pasta_mes + "/Extratos BTG/"
path_pasta_itau = path_pasta_mes + '/Extratos Itaú/'
path_pasta_santander = path_pasta_mes + '/Extratos Santander/'
path_pasta_xp = path_pasta_mes + '/Extratos XP/'

path_pastas = [path_pasta_rda_fi_multi,
               path_pasta_rda_ii,
               path_pasta_rda_prev,
               path_pasta_btg,
               path_pasta_itau,
               path_pasta_santander,
               path_pasta_xp]

arquivos_por_pasta = {}
for pasta in path_pastas:
    arquivos = os.listdir(pasta)
    arquivos_por_pasta[pasta] = arquivos
    #print(arquivos)


nomes = ['cliente_2', 'cliente_1', 'cliente_3', 'cliente_4', 'cliente_5', 'fundo_1', 'fundo_2', 'fundo_3']
arquivos_dict = {}

#gera dicionario com chave igual ao nome da pessoa e valor igual a uma lista
#com o nome de todos os arquivos que tenham em seu conteúdo a chave em questão (nome)
for path in arquivos_por_pasta:
    for extrato_file in range(len(arquivos_por_pasta[path])):
        for nome in nomes:
            if nome in arquivos_por_pasta[path][extrato_file].upper():
                lista_nome = nome
                if lista_nome not in arquivos_dict:
                    arquivos_dict[lista_nome] = []
                #cada chave do dicionario representa uma pessoa
                #cada chave recebe uma lista com todos os path dos arquivos
                #relacionados àquela pessoa
                arquivos_dict[lista_nome].append(path + arquivos_por_pasta[path][extrato_file])

root.destroy()

#selecionar arquivos que serão lidos
root2 = tk.Tk()
contas_menu = ['cliente_2 Itaú',
            'cliente_2 Santander',
            'cliente_1 BTG',
            'cliente_1 Santander',
            'cliente_1 XP',
            'cliente_3 Santander',
            'cliente_5 Santander',
            'cliente_4 Santander',
            'fundo_1',
            'RDA FI ',
            'fundo_3']

vars = []
leitura_arquivos = {}
margem = 10
for i, conta in enumerate(contas_menu):
    if i % 2 == 0:
        color = "#E0E0E0"
    else:
        color = "white"
    tk.Label(root2, text=conta, bg=color, font=("Helvetica", 12), padx=margem, pady=margem).grid(row=i, column=0, sticky="nswe")
    var = tk.IntVar()
    tk.Radiobutton(root2, text='Sim', variable=var, value=1, bg=color, font=("Helvetica", 12), padx=margem, pady=margem).grid(row=i, column=1, sticky="nswe")
    tk.Radiobutton(root2, text='Não', variable=var, value=0, bg=color, font=("Helvetica", 12), padx=margem, pady=margem).grid(row=i, column=2, sticky="nswe")
    vars.append(var)
    root2.rowconfigure(i, weight=1)

root2.columnconfigure(0, weight=1)
root2.columnconfigure(1, weight=1)
root2.columnconfigure(2, weight=1)

tk.Button(root2, text='OK', command=on_submit, font=("Helvetica", 12)).grid(row=len(contas_menu), column=1, pady=margem, padx=margem)
messagebox.showinfo("Atualizar contas", "Escolha sim ou não para cada conta bancária")
root2.mainloop()

root2.destroy()

#extrato btg cliente_1
if leitura_arquivos["cliente_1 BTG"] == 1:
    for extrato in arquivos_dict['cliente_1']:
        if "Extratos BTG" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[2]).extract_text()).splitlines()
    for linha in range(0, len(arquivo)):
        if (arquivo[linha][0:5]).upper() == "TOTAL":
            saldo_btg_olgah = float(((arquivo[linha].split()[3]).replace(".", "")).replace(",", "."))
            break
else:
    saldo_btg_olgah = 0

desconto_itau = 0
#extrato itau cliente_2
if leitura_arquivos["cliente_2 Itaú"] == 1:
    for extrato in arquivos_dict['cliente_2']:
        if "Extratos Itaú" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[3]).extract_text()).splitlines()
    #saldo total, sem o desconto
    for linha in range(0, len(arquivo)):
        if (arquivo[linha][0:16]).upper() == "TOTAL RENDA FIXA":
            saldo_bruto = float(((arquivo[linha].split()[4]).replace(".", "")).replace(",", "."))
            break
    #definir valor do desconto e o saldo final
    for linha in range(0, len(arquivo)):
        if arquivo[linha] == "53565 -ITKINEA RF" and arquivo[linha + 1] == "ABS LPF":
            desconto_itau = float(((arquivo[linha + 2].split()[3]).replace(".", "")).replace(",", "."))
            break
    saldo_itau_dilson = saldo_bruto - desconto_itau
else:
    saldo_itau_dilson = 0

#extrato cliente_1 xp
if leitura_arquivos["cliente_1 XP"] == 1:
    for extrato in arquivos_dict['cliente_1']:
        if "Extratos XP" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[0]).extract_text()).splitlines()
    for linha in range(0, len(arquivo)):
        if arquivo[linha] == 'EVOLUÇÃO PATRIMONIAL' and arquivo[linha + 1] == "POSIÇÃO":
            saldo_xp_olgah = float((((arquivo[linha + 2]).replace("R$", "")).replace(".", "")).replace(",", "."))
            break
else:
    saldo_xp_olgah = 0

#extrato cliente_1 santander
if leitura_arquivos["cliente_1 Santander"] == 1:
    for extrato in arquivos_dict['cliente_1']:
        if "Extratos Santander" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[4]).extract_text()).splitlines()
    for linha in range(0, len(arquivo)):
        if (arquivo[linha][0:5]).upper() == "TOTAL" and len(arquivo[linha + 1].split()) == 5:
            saldo_santander_olgah = float((((arquivo[linha + 1]).split()[0]).replace(".", "")).replace(",", "."))
            break
        elif (arquivo[linha][0:5]).upper() == "TOTAL" and len(arquivo[linha + 1].split()) == 6:
            saldo_santander_olgah = float((((arquivo[linha + 1]).split()[1]).replace(".", "")).replace(",", "."))
            break
else:
    saldo_santander_olgah = 0

#extrato cliente_5 santander
if leitura_arquivos["cliente_5 Santander"] == 1:
    for extrato in arquivos_dict['cliente_5']:
        if "Extratos Santander" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[4]).extract_text()).splitlines()
    for linha in range(0, len(arquivo)):
        if (arquivo[linha][0:5]).upper() == "TOTAL" and len(arquivo[linha + 1].split()) == 5:
            saldo_santander_cliente_5 = float((((arquivo[linha + 1]).split()[0]).replace(".", "")).replace(",", "."))
            break
        elif (arquivo[linha][0:5]).upper() == "TOTAL" and len(arquivo[linha + 1].split()) == 6:
            saldo_santander_cliente_5 = float((((arquivo[linha + 1]).split()[1]).replace(".", "")).replace(",", "."))
            break
else:
    saldo_santander_cliente_5 = 0

#extrato cliente_3 santander
if leitura_arquivos["cliente_4 Santander"] == 1:
    for extrato in arquivos_dict['cliente_3']:
        if "Extratos Santander" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[4]).extract_text()).splitlines()
    for linha in range(0, len(arquivo)):
        if (arquivo[linha][0:5]).upper() == "TOTAL" and len(arquivo[linha + 1].split()) == 5:
            saldo_santander_tapajos = float((((arquivo[linha + 1]).split()[0]).replace(".", "")).replace(",", "."))
            break
        elif (arquivo[linha][0:5]).upper() == "TOTAL" and len(arquivo[linha + 1].split()) == 6:
            saldo_santander_tapajos = float((((arquivo[linha + 1]).split()[1]).replace(".", "")).replace(",", "."))
            break
else:
    saldo_santander_tapajos = 0

#extrato monica santander
if leitura_arquivos["cliente_3 Santander"] == 1:
    for extrato in arquivos_dict['cliente_4']:
        if "Extratos Santander" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[4]).extract_text()).splitlines()
    for linha in range(0, len(arquivo)):
        if (arquivo[linha][0:5]).upper() == "TOTAL" and len(arquivo[linha + 1].split()) == 5:
            saldo_santander_monica = float((((arquivo[linha + 1]).split()[0]).replace(".", "")).replace(",", "."))
            break
        elif (arquivo[linha][0:5]).upper() == "TOTAL" and len(arquivo[linha + 1].split()) == 6:
            saldo_santander_monica = float((((arquivo[linha + 1]).split()[1]).replace(".", "")).replace(",", "."))
            break
else:
    saldo_santander_monica = 0

#extrato cliente_2 santander
if leitura_arquivos["cliente_2 Santander"] == 1:
    for extrato in arquivos_dict['cliente_2']:
        if "Extratos Santander" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[4]).extract_text()).splitlines()
    for linha in range(0, len(arquivo)):
        if (arquivo[linha][0:5]).upper() == "TOTAL" and len(arquivo[linha + 1].split()) == 5:
            saldo_santander_bruto = float((((arquivo[linha + 1]).split()[0]).replace(".", "")).replace(",", "."))
            break
        elif (arquivo[linha][0:5]).upper() == "TOTAL" and len(arquivo[linha + 1].split()) == 6:
            saldo_santander_bruto = float((((arquivo[linha + 1]).split()[1]).replace(".", "")).replace(",", "."))
            break
    #definir desconto e saldo final
    arquivo = ((PdfReader(pdf_path).pages[5]).extract_text()).splitlines()
    for linha in range(0, len(arquivo)):
        lista_linha = arquivo[linha].split()
        for lista in range(0, len(lista_linha)):
            if "10/08/2022" in lista_linha and "11/08/2025" in lista_linha:
                desconto_santander = float((lista_linha[9].replace(".", "")).replace(",", "."))
                break
    saldo_santander_dilson = saldo_santander_bruto - desconto_santander
else:
    saldo_santander_dilson = 0

#extrato fundo_1
if leitura_arquivos["fundo_1"] == 1:
    for extrato in arquivos_dict['fundo_1']:
        if "Extrato fundo_1" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[0]).extract_text()).splitlines()
    for linha in range(0, len(arquivo)):
        if "PÁGINA" in arquivo[linha].upper() and "fundo_1 FIM" in arquivo[linha].upper() and "RENTABILIDADES" in arquivo[linha + 1].upper():
            saldo_rda_liffe_prev = float((arquivo[linha + 4].split()[1]).replace(",", ""))
            break
else:
    saldo_rda_liffe_prev = 0

#extrato RDA FI Multim
if leitura_arquivos["RDA FI "] == 1:
    for extrato in arquivos_dict['fundo_2']:
        if "Extrato RDA FI Multim" in extrato and ".pdf" in extrato:
            pdf_path = extrato
            arquivo = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
    dados = []
    # criar lista com os dados de cada tabela
    for c, tabela in enumerate(arquivo):
        dados.append(tabela.values)
    arquivo = []
    continua = True
    arquivo_conta = []
    cont = 0  # validar saida do ultimo laço for

    for tabela in range(0, len(dados)):
        if cont == 0:
            for linha in range(2, len(dados[tabela])):
                if "Contas a Pagar/Receber" not in dados[tabela][linha] and continua == True:
                    lista_array = np.array(dados[tabela][linha])
                    arquivo.append(lista_array.tolist())
                    lista_array = ""
                elif "Contas a Pagar/Receber" in dados[tabela][linha]:
                    continua = False
                    lista_array = np.array(dados[tabela][linha])
                    arquivo_conta.append(lista_array.tolist())
                    lista_array = ""
                elif continua == False and "Rentabilidade" not in dados[tabela][linha]:
                    lista_array = np.array(dados[tabela][linha])
                    arquivo_conta.append(lista_array.tolist())
                    lista_array = ""
                if "Rentabilidade" in dados[tabela][linha]:
                    cont = 1
                    break
        else:
            break

    arquivo_resumo = []
    saldo_rda_fi_multim = 0
    for sublist in arquivo:
        if len(sublist) >= 11:
            if str(sublist[10]).count(",") > 0:
                sublist[10] = float(str(sublist[10]).replace(",", ""))
                saldo_rda_fi_multim += sublist[10]
            else:
                if str(sublist[16]).count(",") > 0:
                    sublist[16] = float(str(sublist[16]).replace(",", ""))
                    saldo_rda_fi_multim += sublist[16]
            arquivo_resumo.append([sublist[0], sublist[1], sublist[10], sublist[16]])

    arquivo_extra = []
    dados = []
    for lista in range(len(arquivo_conta)):
        if "Total Geral" in arquivo_conta[lista]:
            dados.append(arquivo_conta[lista][1])
            dados.append(float(str(arquivo_conta[lista][3]).replace("(", "").replace(")", "").replace(",", "")))
            saldo_rda_fi_multim += (-1 * float(str(arquivo_conta[lista][3]).replace("(", "").replace(")", "").replace(",", "")))
            arquivo_extra.append(dados)
            dados = []
        elif 'Saldo em Tesouraria' in arquivo_conta[lista]:
            dados.append(arquivo_conta[lista][0])
            dados.append(float(str(arquivo_conta[lista][1]).replace(",", "")))
            saldo_rda_fi_multim += float(str(arquivo_conta[lista][1]).replace(",", ""))
            arquivo_extra.append(dados)
            dados = []

    continua = True
    ultima_linha = 0
    for lista in range(len(arquivo)):
        if continua == True:
            for elemento in range(1, 3):
                class_rda_multi.cell(row=lista + 2, column=elemento).value = arquivo_resumo[lista][elemento]
                linha_final = lista + 2
                if arquivo_resumo[lista][elemento - 1] == "Código":
                    continua = False
                    break
        else:
            if arquivo_resumo[lista][0] and arquivo_resumo[lista][3]:
                class_rda_multi.cell(row=linha_final + 1, column=1).value = arquivo_resumo[lista][0]
                class_rda_multi.cell(row=linha_final + 1, column=2).value = arquivo_resumo[lista][3]
                linha_final += 1
                ultima_linha = linha_final + 1

    class_rda_multi.cell(row=ultima_linha, column=1).value = "Contas a pagar/receber"
    class_rda_multi.cell(row=ultima_linha, column=2).value = -1 * arquivo_extra[0][1]
    class_rda_multi.cell(row=ultima_linha + 1, column=1).value = arquivo_extra[1][0]
    class_rda_multi.cell(row=ultima_linha + 1, column=2).value = arquivo_extra[1][1]
else:
    saldo_rda_fi_multim = 0

#extrato fundo_3
if leitura_arquivos["fundo_3"] == 1:
    for extrato in arquivos_dict['fundo_3']:
        if "Extrato fundo_3" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[0]).extract_text()).splitlines()
    for linha in range(0, len(arquivo)):
        if "PATRIMÔNIO LÍQUIDO" in arquivo[linha].upper():
            if len(arquivo[linha + 1]) > 6:
                saldo_rda_ii_liffe = float(((arquivo[linha + 1].replace(".", "")).replace(",", ".")))
            else:
                saldo_rda_ii_liffe = float((arquivo[linha + 1].replace(",", ".")))
            break
else:
    saldo_rda_ii_liffe = 0

#escrever saldos PDFs na planilha
contas = {"cliente_1 BTG": saldo_btg_olgah,
        "cliente_2 BTG": 0,
        "fundo_1": saldo_rda_liffe_prev,
        "fundo_2": saldo_rda_fi_multim,
        "fundo_3": saldo_rda_ii_liffe,
        "cliente_1 Santander": saldo_santander_olgah,
        "cliente_2 Santander": saldo_santander_dilson,
        "cliente_3 Santander": saldo_santander_monica,
        "cliente_5 Santander": saldo_santander_cliente_5,
        "cliente_4 Santander": saldo_santander_tapajos,
        "cliente_1 Itaú": 0,
        "cliente_2 Itaú": saldo_itau_dilson,
        "cliente_1 XP": saldo_xp_olgah}


for conta in range(2, len(contas) + 3):
    for key, saldo in contas.items():
        if checks.cell(row=conta, column=1).value == key:
            checks.cell(row=conta, column=2).value = saldo
            break


#extrato fundo_3 - lista de ativos
if leitura_arquivos["fundo_3"] == 1:
    for extrato in arquivos_dict['fundo_3']:
        if "Extrato fundo_3" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)

    dados = []
    continua = True
    #tabela = conteudo tabela
    # criar lista com os dados de cada tabela
    for c, tabela in enumerate(arquivo):
        dados.append(tabela.values)

    arquivo = []

    for tabela in range(1, len(dados)):
        if continua == False:
            break
        for linha in range(0, len(dados[tabela])):
            if dados[tabela][linha][0] == "Contas a pagar e receber":
                lista_array = np.array(dados[tabela][linha])
                arquivo.append(lista_array.tolist())
                continua = False
                break
            elif len(dados[tabela][linha]) > 7:
                lista_array = np.array(dados[tabela][linha])
                arquivo.append(lista_array.tolist())

    #escrever ativos na planilha aba Classificação RDA II
    for lista in range(0, len(arquivo)):#linha
        for elemento in range(0, 2):#coluna
            class_rda_ii.cell(row=lista+1, column=elemento+1).value = formatar(arquivo[lista][elemento])


#extrato fundo_1 - lista de ativos
if leitura_arquivos["fundo_1"] == 1:
    for extrato in arquivos_dict['fundo_1']:
        if "Extrato fundo_1" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)

    dados = []
    continua = True
    #tabela = conteudo tabela
    # criar lista com os dados de cada tabela
    for c, tabela in enumerate(arquivo):
        dados.append(tabela.values)
    arquivo = []


    for tabela in range(2, len(dados)):
        if continua == False:
            break
        for linha in range(0, len(dados[tabela])):
            if "REPASSE PI" in dados[tabela][linha]:
                lista_array = np.array(dados[tabela][linha])
                arquivo.append(lista_array.tolist())
                continua = False
                break
            else:
                lista_array = np.array(dados[tabela][linha])
                arquivo.append(lista_array.tolist())


    for c in range(0, len(arquivo)):
        if str(arquivo[c][0]).count("- ") > 0:
            arquivo[c][0] = str(arquivo[c][0]).split("- ")[1]

    #escrever ativos na planilha aba Classificação fundo_1
    #apagar "nan" para salvar na coluna correta os valores a serem somados
    for sublist in arquivo:
        for i, elem in enumerate(sublist):
            if pd.isna(elem):
                sublist.pop(i)

    for lista in range(0, len(arquivo)):#linha
        for elemento in range(0, 3):#coluna
            if (str(arquivo[lista][elemento])).count(".") > 0:
                class_rda_prev.cell(row=lista + 2, column=elemento + 1).value = float((str(arquivo[lista][elemento]).replace(",", "")))
            elif arquivo[lista][0] != 'Ativo':
                class_rda_prev.cell(row=lista + 2, column=elemento + 1).value = arquivo[lista][elemento]


#ler movimentações santander cliente_2
if leitura_arquivos["cliente_2 Santander"] == 1:
    for extrato in arquivos_dict['cliente_2']:
        if "Extratos Santander" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[5]).extract_text()).splitlines()

    dados = []

    for linha in range(0, len(arquivo)):
        if len(arquivo[linha].split()) > 10:
            dados.append(arquivo[linha].split())

    arquivo = []


    for lista in dados:
        if "10/08/2022" not in lista and len(lista) < 20:
            if "DIPB" in lista :
                lista.remove("DIPB")
            if "DI" in lista:
                lista.remove("DI")
            if "PB" in lista:
                lista.remove("PB")
            lista[3] = lista[3] + " " + lista[4] + "%"
            del lista[4]
            if len(lista) == 13:
                lista[12] = "Santander"
            if lista[6].count(".") > 0:
                lista[6] = float((lista[6].replace(".", "")).replace(",", "."))
            if len(lista) == 12:
                aux = lista[4]
                lista[4] = lista[4][:10]
                lista.insert(5, aux[10:])
                lista[12] = "Santander"
            if str(lista[6]).count(",") > 0:
                lista[6] = float(str(lista[6]).replace(".", "").replace(",", "."))
            else:
                lista[6] = float(lista[6])
            if str(lista[9]).count(",") > 0:
                lista[9] = float(str(lista[9]).replace(".", "").replace(",", "."))
            else:
                lista[9] = float(lista[9])
            arquivo.append(lista)


    for linha in range(0, len(arquivo)):
        for coluna in range(0, len(arquivo[linha])):
            santander.cell(row=linha + 2, column=coluna + 1).value = arquivo[linha][coluna]

#ler movimentações santander cliente_5
if leitura_arquivos["cliente_5 Santander"] == 1:
    for extrato in arquivos_dict['cliente_5']:
        if "Extratos Santander" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[5]).extract_text()).splitlines()
    dados = []

    for linha in range(0, len(arquivo)):
        if len(arquivo[linha].split()) > 10:
            dados.append(arquivo[linha].split())

    arquivo = []
    for lista in dados:
        if len(lista) < 20:
            if "DIPB" in lista:
                lista.remove("DIPB")
            if "DI" in lista:
                lista.remove("DI")
            if "PB" in lista:
                lista.remove("PB")
            lista[3] = lista[3] + " " + lista[4] + "%"
            del lista[4]
            if len(lista) == 13:
                lista[12] = "Santander"
            if lista[6].count(".") > 0:
                lista[6] = float((lista[6].replace(".", "")).replace(",", "."))
            if len(lista) == 12:
                aux = lista[4]
                lista[4] = lista[4][:10]
                lista.insert(5, aux[10:])
                lista[12] = "Santander"
            if str(lista[6]).count(",") > 0:
                lista[6] = float(str(lista[6]).replace(".", "").replace(",", "."))
            else:
                lista[6] = float(lista[6])
            if str(lista[9]).count(",") > 0:
                lista[9] = float(str(lista[9]).replace(".", "").replace(",", "."))
            else:
                lista[9] = float(lista[9])
            arquivo.append(lista)



    for linha in range(0, len(arquivo)):
        for coluna in range(0, len(arquivo[linha])):
            santander.cell(row=linha + 10, column=coluna + 1).value = arquivo[linha][coluna]


#ler movimentações santander cliente_3
if leitura_arquivos["cliente_4 Santander"] == 1:
    for extrato in arquivos_dict['cliente_3']:
        if "Extratos Santander" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[5]).extract_text()).splitlines()

    dados = []

    for linha in range(0, len(arquivo)):
        if len(arquivo[linha].split()) > 10:
            dados.append(arquivo[linha].split())

    arquivo = []

    for lista in dados:
        if len(lista) < 20:
            if "DIPB" in lista:
                lista.remove("DIPB")
            if "DI" in lista:
                lista.remove("DI")
            if "PB" in lista:
                lista.remove("PB")
            lista[3] = lista[3] + " " + lista[4] + "%"
            del lista[4]
            if len(lista) == 13:
                lista[12] = "Santander"
            if lista[6].count(".") > 0:
                lista[6] = float((lista[6].replace(".", "")).replace(",", "."))
            if len(lista) == 12:
                aux = lista[4]
                lista[4] = lista[4][:10]
                lista.insert(5, aux[10:])
                lista[12] = "Santander"
            if str(lista[6]).count(",") > 0:
                lista[6] = float(str(lista[6]).replace(".", "").replace(",", "."))
            else:
                lista[6] = float(lista[6])
            if str(lista[9]).count(",") > 0:
                lista[9] = float(str(lista[9]).replace(".", "").replace(",", "."))
            else:
                lista[9] = float(lista[9])
            arquivo.append(lista)



    for linha in range(0, len(arquivo)):
        for coluna in range(0, len(arquivo[linha])):
            santander.cell(row=linha + 18, column=coluna + 1).value = arquivo[linha][coluna]


#ler movimentações santander cliente_1
if leitura_arquivos["cliente_1 Santander"] == 1:
    for extrato in arquivos_dict['cliente_1']:
        if "Extratos Santander" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[5]).extract_text()).splitlines()

    dados = []

    for linha in range(0, len(arquivo)):
        if len(arquivo[linha].split()) > 10:
            dados.append(arquivo[linha].split())

    arquivo = []

    for lista in dados:
        if len(lista) < 20:
            if "DIPB" in lista:
                lista.remove("DIPB")
            if "DI" in lista:
                lista.remove("DI")
            if "PB" in lista:
                lista.remove("PB")
            lista[3] = lista[3] + " " + lista[4] + "%"
            del lista[4]
            if len(lista) == 13:
                lista[12] = "Santander"
            if lista[6].count(".") > 0:
                lista[6] = float((lista[6].replace(".", "")).replace(",", "."))
            if len(lista) == 12:
                aux = lista[4]
                lista[4] = lista[4][:10]
                lista.insert(5, aux[10:])
                lista[12] = "Santander"
            if str(lista[6]).count(",") > 0:
                lista[6] = float(str(lista[6]).replace(".", "").replace(",", "."))
            else:
                lista[6] = float(lista[6])
            if str(lista[9]).count(",") > 0:
                lista[9] = float(str(lista[9]).replace(".", "").replace(",", "."))
            else:
                lista[9] = float(lista[9])
            arquivo.append(lista)


    for linha in range(0, len(arquivo)):
        for coluna in range(0, len(arquivo[linha])):
            santander.cell(row=linha + 26, column=coluna + 1).value = arquivo[linha][coluna]


#ler movimentações santander cliente_3
if leitura_arquivos["cliente_3 Santander"] == 1:
    for extrato in arquivos_dict['cliente_4']:
        if "Extratos Santander" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[5]).extract_text()).splitlines()

    dados = []

    for linha in range(0, len(arquivo)):
        if len(arquivo[linha].split()) > 10:
            dados.append(arquivo[linha].split())

    arquivo = []

    for lista in dados:
        if len(lista) < 20:
            if "DIPB" in lista:
                lista.remove("DIPB")
            if "DI" in lista:
                lista.remove("DI")
            if "PB" in lista:
                lista.remove("PB")
            lista[3] = lista[3] + " " + lista[4] + "%"
            del lista[4]
            if len(lista) == 13:
                lista[12] = "Santander"
            if lista[6].count(".") > 0:
                lista[6] = float((lista[6].replace(".", "")).replace(",", "."))
            if len(lista) == 12:
                aux = lista[4]
                lista[4] = lista[4][:10]
                lista.insert(5, aux[10:])
                lista[12] = "Santander"
            if str(lista[6]).count(",") > 0:
                lista[6] = float(str(lista[6]).replace(".", "").replace(",", "."))
            else:
                lista[6] = float(lista[6])
            if str(lista[9]).count(",") > 0:
                lista[9] = float(str(lista[9]).replace(".", "").replace(",", "."))
            else:
                lista[9] = float(lista[9])
            arquivo.append(lista)

    for linha in range(0, len(arquivo)):
        for coluna in range(0, len(arquivo[linha])):
            santander.cell(row=linha + 34, column=coluna + 1).value = arquivo[linha][coluna]



#ler movimentações BTG cliente_1
if leitura_arquivos["cliente_1 BTG"] == 1:
    for extrato in arquivos_dict['cliente_1']:
        if "Extratos BTG" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)

    dados = []
    continua = True

    # criar lista com os dados de cada tabela
    for c, tabela in enumerate(arquivo):
        dados.append(tabela.values)
    arquivo = []

    for tabela in range(2, len(dados)):
        for linha in range(0, len(dados[tabela])):
            lista_array = np.array(dados[tabela][linha])
            arquivo.append(lista_array.tolist())

    dados_individual = []
    dados_geral = []

    for sublist in arquivo:
        if sublist[0][:5] == "BANCO":
            break
        for i, elem in enumerate(sublist):
            if pd.isna(elem) == False:
                if "Ativo" != sublist[0][:5] and "Quantidade" != sublist[0][:10] and "Detalhamento" != sublist[0][:12] and "IOF" != sublist[0][:3] and "Total" != sublist[0][:5] and ":" != sublist[0][:1]:
                    dados_individual.append(elem)
                    add = True
                else:
                    add = False
                    break
        if add == True:
            dados_geral.append(dados_individual)
            dados_individual = []

    lista_individual = []
    lista_final = []
    for e, lista in enumerate(dados_geral):
        if e == 0:
            lista_individual.extend(lista)
        elif e > 0:
            if lista[0][:4] == "CDB-":
                lista_final.append(lista_individual)
                lista_individual = []
                lista_individual.extend(lista)
            else:
                lista_individual.extend(lista)
    lista_final.append(lista_individual)


    for sub in range(len(lista_final)):
        for elemento in range(len(lista_final[sub])):
            if str(lista_final[sub][elemento]).count("% ") > 0:
                lista_final[sub][elemento] = str(lista_final[sub][elemento]).split()[0]
            if str(lista_final[sub][elemento]).count(" ") > 0:
                if str(lista_final[sub][elemento]).split()[0] == "-":
                    lista_final[sub][elemento] = str(lista_final[sub][elemento]).split()[1]
                else:
                    lista_final[sub][elemento] = str(lista_final[sub][elemento]).split()[0]


    arquivo = []
    for lista in lista_final:
        dados = []
        dados.append(lista[0])
        ano = "20"
        ano += lista[2][-2:]
        dia_mes = lista[2][:-2]
        dados.append(dia_mes + ano)
        ano = "20"
        ano += lista[1][-2:]
        dia_mes = lista[1][:-2]
        dados.append(dia_mes + ano)
        dados.append(str(lista[4]) + " CDI")
        dados.append(float(str(lista[8]).replace(".", "").replace(",", ".")))
        dados.append(float(str(lista[-1]).replace(".", "").replace(",", ".")))
        arquivo.append(dados)

    for linha in range(0, len(arquivo)):
        for coluna in range(0, len(arquivo[linha])):
            btg.cell(row=linha + 2, column=coluna + 1).value = arquivo[linha][coluna]


#extrato itau cliente_2
if leitura_arquivos["cliente_2 Itaú"] == 1:
    for extrato in arquivos_dict['cliente_2']:
        if "Extratos Itaú" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[3]).extract_text()).splitlines()

    lista = []
    lista_final = []
    for linha in range(len(arquivo)):
        if "Pós-fixado" in arquivo[linha]:
            indice = linha + 1
            for linha_filtro in range(len(arquivo) - indice):
                if "Saldo emconta" not in arquivo[indice + linha_filtro] and "Prefixado" not in arquivo[indice + linha_filtro] and "Total Renda Fixa" not in arquivo[indice + linha_filtro] and "Portfólio" not in arquivo[indice + linha_filtro]:
                    lista.append(arquivo[indice + linha_filtro])


    cont = 0
    arquivo = []
    for elemento in lista:
        cont += 1
        if cont % 3 != 0:
            lista_final.append(elemento)
        else:
            lista_final.extend(elemento.split())
            arquivo.append(lista_final)
            lista_final = []

    for apagar in range(len(arquivo)):
        if arquivo[apagar][0] == '53565 -ITKINEA RF':
            arquivo.pop(apagar)
            break


    ano = "20"
    for sublista in arquivo:
        if sublista[2] == "PRE":
            sublista[1] = sublista[2][:3]
            sublista[2] = sublista[3] + "%"
            sublista.pop(3)
        if sublista[2][:3] == "CDI" and len(sublista[2]) > 3:
            sublista[1] = sublista[2][:3]
            sublista[2] = sublista[2][3:] + "% "
        elif sublista[2][:3] == "CDI" and len(sublista[2]) == 3:
            sublista[1] = sublista[2][:3]
            sublista[2] = "100%"
        dia_mes = sublista[3][:6]
        sublista[3] = dia_mes + ano + sublista[3][-2:]
        dia_mes = sublista[4][:6]
        sublista[4] = dia_mes + ano + sublista[4][-2:]
        sublista[8] = float(str(sublista[8]).replace(".", "").replace(",", "."))
        sublista[11] = float(str(sublista[11]).replace(".", "").replace(",", "."))

    lista_apagar = [0, 1, 2, 3, 4, 8, 11]
    #apaga elementos desnecessários
    for i in range(len(arquivo)):
        arquivo[i] = [arquivo[i][j] for j in lista_apagar]
        if arquivo[i][1] == "PRE":
            arquivo[i][2] = arquivo[i][2].replace(",", ".")

    for linha in range(0, len(arquivo)):
        for coluna in range(0, len(arquivo[linha])):
            itau.cell(row=linha + 2, column=coluna + 1).value = arquivo[linha][coluna]


#extrato cliente_1 xp
if leitura_arquivos["cliente_1 XP"] == 1:
    for extrato in arquivos_dict['cliente_1']:
        if "Extratos XP" in extrato and ".pdf" in extrato:
            pdf_path = extrato
    arquivo = ((PdfReader(pdf_path).pages[0]).extract_text()).splitlines()

    dados = []
    lista_final = []
    continua = True
    cont = 0

    for i in range(len(arquivo)):
        if continua == False:
            break
        if arquivo[i] == "LÍQUIDO":
           inicio = i + 1
           for lista in range(inicio, len(arquivo)):
                if arquivo[lista] != 'DISTRIBUIÇÃO DE LIQUIDEZ DA CARTEIRA':
                    if cont % 11 != 0 or cont == 0:
                        dados.append(arquivo[lista])
                    else:
                        dados.append(arquivo[lista])
                        lista_final.append(dados)
                        dados = []
                    cont += 1
                else:
                    continua = False
                    break

    lista_apagar = [0, 3, 5, 6, 10, 11]
    arquivo = []
    #apaga elementos desnecessários
    for i in range(len(lista_final)):
        lista_final[i] = [lista_final[i][j] for j in lista_apagar]

    for sublista in lista_final:
        sublista[4] = float(sublista[4][3:].replace(".", "").replace(",", "."))
        sublista[5] = float(sublista[5][3:].replace(".", "").replace(",", "."))

    for linha in range(0, len(lista_final)):
        for coluna in range(0, len(lista_final[linha])):
            xp.cell(row=linha + 2, column=coluna + 1).value = lista_final[linha][coluna]


#extrair dados do quantum com leitura de imagem
leitura_saldo_quantum = pyautogui.confirm('Deseja ler o extrato da Quantum?', buttons=['Sim', 'Não'], title='Quantum')
if leitura_saldo_quantum == "Sim":
    #data base do saldo
    root_data = tk.Tk()
    root_data.withdraw()
    validacao_until = False
    while validacao_until == False:
        data = simpledialog.askstring("Data base", "Digite a data base dd/mm/aaaa")
        if str(data).count("/") == 2:
            input_until_validacao = str(data).split("/")
            if len(input_until_validacao) != 3:
                validacao_until = False
            else:
                if len(input_until_validacao[0]) == 2 and len(input_until_validacao[1]) == 2 and len(input_until_validacao[2]) == 4:
                    validacao_until = True
    root_data.destroy()

    #ocultar navegador do usuário
    opcao = webdriver.ChromeOptions()
    #opcao.add_argument("--headless")
    #criar navegador inflation
    navegador = webdriver.Chrome(options=opcao)

    # acessar site
    navegador.get('https://www.quantumaxis.com.br')
    navegador.implicitly_wait(5)

    login = navegador.find_element(By.XPATH, '//*[@id="campoLogin"]')
    senha = navegador.find_element(By.XPATH, '//*[@id="campoSenha"]')
    login.send_keys('login') #digita login
    senha.send_keys('senha') #digita senha

    navegador.find_element(By.XPATH, '//*[@id="botaoEntrar"]').click() #clica em entrar
    portifolio = navegador.find_element(By.XPATH, '//*[@id="itens"]/li[5]')
    mover = ActionChains(navegador)
    mover.move_to_element(portifolio) #move ate menu portifolio
    mover.perform()

    monitoramento = navegador.find_element(By.XPATH, '//*[@id="linkMonitoramento"]')

    lista_ativos = ['nome_cliente (BTG)', 'fundo_1 FI ', 'RDA FI ', 'fundo_3 FI ', 'nome_cliente (Santander)', 'nome_cliente ita', 'nome_cliente - xp']

    navegador.find_element(By.XPATH, '//*[@id="linkMonitoramento"]').click() #clica monitoramento
    sleep(8)
    saldo_lista = []
    pyautogui.click(157, 521) #espaço para digitar nome ativo
    for ativo in lista_ativos:
        pyautogui.write(ativo)
        sleep(3)
        pyautogui.click(162, 578) #clica no ativo selecionado
        pyautogui.click(206, 669) #clica report
        sleep(4)
        pyautogui.click(1295, 363) #data
        pyautogui.press("backspace", presses=15)
        pyautogui.write(data)
        sleep(2)
        screenshot = pyautogui.screenshot()
        screenshot.save(f'{ativo}.png')
        # Open image
        img = cv2.imread(f"{ativo}.png")
        # Convert the image to grayscale
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # Apply Otsu thresholding
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
        # Pass the thresholded image to Tesseract OCR
        text = str(pytesseract.image_to_string(thresh, lang='eng')).splitlines()
        #selecionar saldo
        for linha in text:
            if "Saldo Total:" in linha:
                for letra in range(len(linha)):
                    if linha[letra].isdigit():
                        saldo_lista.append(float(linha[letra:letra+14].replace(".", "").replace(",", ".")))
                        break
        pyautogui.click(393, 360)
        sleep(1)
        pyautogui.click(441, 524)
        pyautogui.press("backspace", presses=40)
        sleep(2)

    for saldo in range(len(saldo_lista)):
        quantum.cell(row=saldo + 2, column=2).value = saldo_lista[saldo]

    navegador.quit()

planilha.save(planilha_path)

pyautogui.alert('Planilha atualizada.', button='OK', title='Atualizada')

os.startfile(planilha_path)


#pyinstaller --onefile --add-binary  "C:\Users\Lenovo\Documents\Projetos\nome_cliente\venv\Lib\site-packages\tabula\tabula-1.0.5-jar-with-dependencies.jar;/.tabula/" Validacao_saldos.py
#https://github.com/pyinstaller/pyinstaller/issues/5298
'''
You just need to add the jar file as a data file. It's easiest to do this with the spec file.
Edit the .spec file that should have been generated automatically.
Put at the top:
from PyInstaller.utils.hooks import collect_data_files 
And replace the datas=[] with:
datas=collect_data_files("tabula")
Then rebuild using
pyinstaller Validacao_saldos.spec
from now on.
'''
#https://github.com/UB-Mannheim/tesseract/wiki
#baixar tesseract











